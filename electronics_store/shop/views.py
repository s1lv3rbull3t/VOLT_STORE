from django.contrib import messages
from django.conf import settings
from collections import defaultdict
from django.core.mail import send_mail
from django.core.signing import BadSignature, SignatureExpired, TimestampSigner, dumps, loads
from django.contrib.auth.hashers import make_password, check_password
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.db import transaction
from django.db.models import Avg, Sum, F, ExpressionWrapper, DecimalField, Min, Max
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.decorators.http import require_POST
import re
from urllib.parse import urlencode
from datetime import timedelta
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import (
    Product,
    Category,
    User,
    City,
    Address,
    Order,
    OrderStatus,
    CartItem,
    OrderItem,
    Review,
    CategoryAttributeMap,
    AttributeDefinition,
    AttributeOption,
    ProductAttributeValue,
)
from .services import (
    UserService,
    CategoryService,
    ProductService,
    CartService
)

REGISTER_LINK_TTL_SECONDS = 60 * 60
REGISTER_ACTIVATION_SALT = 'shop.registration.activate'
PROFILE_EMAIL_CHANGE_LINK_TTL_SECONDS = 60 * 60
PROFILE_EMAIL_CHANGE_SALT = 'shop.profile.email.change'


def normalize_phone_number(raw_phone):
    if not raw_phone:
        return None
    digits = re.sub(r'\D', '', raw_phone)
    if not digits:
        return None
    if len(digits) > 15:
        return None
    return digits

def index(request):
    categories = CategoryService.get_root_categories()
    deals = ProductService.get_daily_deals()
    hits = ProductService.get_bestsellers()
    new_arrivals = ProductService.get_new_arrivals()
    
    user = UserService.get_user_from_session(request)
    cart_item_count = CartService.get_cart_item_count(user)

    return render(request, 'shop/index.html', {
        'categories': categories,
        'deals': deals,
        'hits': hits,
        'new_arrivals': new_arrivals,
        'cart_item_count': cart_item_count,
        'current_user': user,
    })



def get_logged_in_user(request):
    return UserService.get_user_from_session(request)


def get_cart_item_count(request):
    user = get_logged_in_user(request)
    return CartService.get_cart_item_count(user)


def get_base_context(request):
    user = get_logged_in_user(request)
    return {
        'cart_item_count': CartService.get_cart_item_count(user),
        'current_user': user,
        'search_query': request.GET.get('q', '').strip(),
    }


def redirect_to_login_with_next(request):
    login_url = reverse('login')
    next_param = urlencode({'next': request.get_full_path()})
    return redirect(f'{login_url}?{next_param}')


def is_admin(user):
    return bool(user and getattr(user, 'role', User.ROLE_BUYER) == User.ROLE_ADMIN)


def can_manage_store(user):
    return is_admin(user)


def get_reviews_label(count):
    last_two = count % 100
    last_one = count % 10
    if 11 <= last_two <= 14:
        form = 'отзывов'
    elif last_one == 1:
        form = 'отзыв'
    elif 2 <= last_one <= 4:
        form = 'отзыва'
    else:
        form = 'отзывов'
    return f'{count} {form}'


def parse_non_negative_int(raw_value):
    if raw_value is None or raw_value == '':
        return None
    try:
        parsed = int(float(raw_value))
    except (TypeError, ValueError):
        return None
    return parsed if parsed >= 0 else None


def parse_float_value(raw_value):
    if raw_value is None or raw_value == '':
        return None
    try:
        return float(raw_value)
    except (TypeError, ValueError):
        normalized = str(raw_value).replace(',', '.').strip()
        try:
            return float(normalized)
        except (TypeError, ValueError):
            return None


def format_day_label(raw_value):
    if hasattr(raw_value, 'strftime'):
        return raw_value.strftime('%d.%m')
    text = str(raw_value or '').strip()
    if not text:
        return ''
    if len(text) >= 10 and text[4] == '-' and text[7] == '-':
        return f'{text[8:10]}.{text[5:7]}'
    return text


def get_review_access(user, product):
    if not user:
        return False, False, False

    has_delivered_purchase = OrderItem.objects.filter(
        id_order__id_user=user,
        id_order__id_status__status_name__iexact='Доставлен',
        id_product=product,
    ).exists()
    already_reviewed = Review.objects.filter(
        id_user=user,
        id_product=product,
    ).exists()
    can_leave_review = has_delivered_purchase and not already_reviewed
    return can_leave_review, has_delivered_purchase, already_reviewed


@require_POST
def add_to_cart(request, product_id):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы добавить товар в корзину, пожалуйста, войдите или зарегистрируйтесь.')
        return redirect_to_login_with_next(request)

    product = get_object_or_404(Product, id_product=product_id)
    if product.stock_quantity <= 0:
        messages.error(request, 'Товар временно отсутствует на складе.')
        next_url = request.POST.get('next') or request.META.get('HTTP_REFERER', '/')
        return redirect(next_url)

    existing_item = CartService.get_cart_item_for_product(user, product)
    previous_quantity = existing_item.quantity if existing_item else 0
    cart_item = CartService.add_product_to_cart(user, product)

    if previous_quantity == (cart_item.quantity if cart_item else 0):
        messages.info(request, 'Достигнуто максимальное количество товара в корзине.')
    else:
        messages.success(request, 'Товар добавлен в корзину.')
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER', '/')
    return redirect(next_url)


def buy_now(request, product_id):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы оформить заказ, пожалуйста, войдите или зарегистрируйтесь.')
        return redirect_to_login_with_next(request)

    product = get_object_or_404(Product, id_product=product_id)
    if product.stock_quantity <= 0:
        messages.error(request, 'Товар временно отсутствует на складе.')
        return redirect('product_detail', product_id=product.id_product)

    existing_item = CartService.get_cart_item_for_product(user, product)
    previous_quantity = existing_item.quantity if existing_item else 0
    cart_item = CartService.add_product_to_cart(user, product)

    if previous_quantity == (cart_item.quantity if cart_item else 0):
        messages.info(request, 'Достигнуто максимальное количество товара в корзине.')

    return redirect('checkout')


@require_POST
def update_cart_item(request, product_id):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы изменить корзину, пожалуйста, войдите.')
        return redirect_to_login_with_next(request)

    product = get_object_or_404(Product, id_product=product_id)
    action = request.POST.get('action', '').strip()
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or '/cart/'

    cart_item = CartService.get_cart_item_for_product(user, product)
    current_quantity = cart_item.quantity if cart_item else 0

    if action == 'increase':
        _, status = CartService.set_cart_item_quantity(user, product, current_quantity + 1)
    elif action == 'decrease':
        _, status = CartService.set_cart_item_quantity(user, product, current_quantity - 1)
    elif action == 'remove':
        _, status = CartService.set_cart_item_quantity(user, product, 0)
    else:
        messages.error(request, 'Неизвестное действие для корзины.')
        return redirect(next_url)

    if status == 'max_reached':
        messages.info(request, 'Достигнут лимит по остатку товара.')
    elif status == 'removed':
        messages.success(request, 'Товар удален из корзины.')
    elif status == 'out_of_stock':
        messages.error(request, 'Товар закончился на складе и удален из корзины.')
    elif status == 'not_in_cart':
        messages.info(request, 'Товар уже отсутствует в корзине.')

    return redirect(next_url)


def cart(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы открыть корзину, пожалуйста, войдите.')
        return redirect_to_login_with_next(request)

    cart_items, total_cost = CartService.get_cart_items_with_totals(user)

    return render(request, 'shop/cart.html', {
        'cart_items': cart_items,
        'total_cost': total_cost,
        'cart_item_count': get_cart_item_count(request),
        'current_user': user,
    })


def checkout(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы оформить заказ, пожалуйста, войдите.')
        return redirect_to_login_with_next(request)

    cart_items, total_cost = CartService.get_cart_items_with_totals(user)
    if not cart_items:
        messages.info(request, 'Корзина пуста. Добавьте товары перед оформлением заказа.')
        return redirect('cart')

    selected_delivery = request.POST.get('delivery_method', 'delivery')
    selected_payment = request.POST.get('payment_method', 'card')

    if request.method == 'POST':
        city_name = request.POST.get('city', '').strip()
        street = request.POST.get('street', '').strip()
        apartment = request.POST.get('apartment', '').strip()

        if selected_delivery == 'delivery' and (not city_name or not street):
            messages.error(request, 'Укажите город и улицу для нового адреса.')
            return render(request, 'shop/checkout.html', {
                'cart_items': cart_items,
                'total_cost': total_cost,
                'cart_item_count': CartService.get_cart_item_count(user),
                'current_user': user,
                'selected_delivery': selected_delivery,
                'selected_payment': selected_payment,
                'form_city': city_name,
                'form_street': street,
                'form_apartment': apartment,
            })

        order_status = OrderStatus.objects.filter(status_name__iexact='Подтвержден').first()
        if not order_status:
            order_status = OrderStatus.objects.create(status_name='Подтвержден')

        with transaction.atomic():
            user_cart = CartService.get_or_create_cart(user)
            checkout_items = list(
                CartItem.objects
                .select_related('id_product')
                .filter(id_cart=user_cart)
            )
            if not checkout_items:
                messages.info(request, 'Корзина пуста. Добавьте товары перед оформлением заказа.')
                return redirect('cart')

            product_ids = [item.id_product_id for item in checkout_items]
            locked_products = {
                product.id_product: product
                for product in Product.objects.select_for_update().filter(id_product__in=product_ids)
            }

            unavailable_items = []
            recalculated_total = 0
            for item in checkout_items:
                product = locked_products.get(item.id_product_id)
                available_qty = product.stock_quantity if product else 0
                if available_qty < item.quantity:
                    unavailable_items.append(
                        f'{item.id_product.product_name} (нужно: {item.quantity}, доступно: {available_qty})'
                    )
                    continue
                recalculated_total += product.price * item.quantity

            if unavailable_items:
                messages.error(
                    request,
                    'Не удалось оформить заказ: некоторые товары закончились или их осталось меньше, чем в корзине. '
                    + 'Проверьте позиции: '
                    + '; '.join(unavailable_items)
                )
                return redirect('cart')

            if selected_delivery == 'delivery':
                city, _ = City.objects.get_or_create(city_name=city_name)
                address = Address.objects.create(
                    id_user=user,
                    id_city=city,
                    street=street,
                    apartment=apartment or None,
                )
            else:
                pickup_city, _ = City.objects.get_or_create(city_name='Самовывоз')
                address = Address.objects.create(
                    id_user=user,
                    id_city=pickup_city,
                    street='Пункт выдачи',
                    apartment=None,
                )

            order = Order.objects.create(
                id_user=user,
                id_address=address,
                total_cost=recalculated_total,
                id_status=order_status,
            )

            OrderItem.objects.bulk_create([
                OrderItem(
                    id_order=order,
                    id_product=item.id_product,
                    quantity=item.quantity,
                )
                for item in checkout_items
            ])

            for item in checkout_items:
                Product.objects.filter(id_product=item.id_product_id).update(
                    stock_quantity=F('stock_quantity') - item.quantity
                )

            CartItem.objects.filter(id_cart=user_cart).delete()

        messages.success(request, 'Заказ успешно оформлен.')
        return redirect('cart')

    return render(request, 'shop/checkout.html', {
        'cart_items': cart_items,
        'total_cost': total_cost,
        'cart_item_count': CartService.get_cart_item_count(user),
        'current_user': user,
        'selected_delivery': selected_delivery,
        'selected_payment': selected_payment,
        'form_city': '',
        'form_street': '',
        'form_apartment': '',
    })


@require_POST
def add_review(request, product_id):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Чтобы оставить отзыв, войдите в аккаунт.')
        return redirect_to_login_with_next(request)

    product = get_object_or_404(Product, id_product=product_id)
    can_leave_review, has_delivered_purchase, already_reviewed = get_review_access(user, product)

    if not has_delivered_purchase:
        messages.error(request, 'Оставить отзыв можно только после доставки заказа с этим товаром.')
        return redirect('product_detail', product_id=product.id_product)

    if already_reviewed or not can_leave_review:
        messages.info(request, 'Вы уже оставили отзыв на этот товар.')
        return redirect('product_detail', product_id=product.id_product)

    rating_raw = request.POST.get('rating', '').strip()
    comment = request.POST.get('comment', '').strip()

    try:
        rating = int(rating_raw)
    except (TypeError, ValueError):
        messages.error(request, 'Укажите корректную оценку от 1 до 5.')
        return redirect('product_detail', product_id=product.id_product)

    if rating < 1 or rating > 5:
        messages.error(request, 'Оценка должна быть от 1 до 5.')
        return redirect('product_detail', product_id=product.id_product)

    Review.objects.create(
        id_user=user,
        id_product=product,
        rating=rating,
        comment=comment or None,
    )
    messages.success(request, 'Спасибо! Ваш отзыв добавлен.')
    return redirect('product_detail', product_id=product.id_product)


def register(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if not all([email, first_name, last_name, password, confirm_password]):
            messages.error(request, 'Пожалуйста, заполните все обязательные поля.')
            return render(request, 'shop/register.html', get_base_context(request))

        if password != confirm_password:
            messages.error(request, 'Пароли не совпадают.')
            return render(request, 'shop/register.html', get_base_context(request))

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Пользователь с таким email уже существует.')
            return render(request, 'shop/register.html', get_base_context(request))

        normalized_phone = normalize_phone_number(phone_number)
        if phone_number and not normalized_phone:
            messages.error(request, 'Телефон должен содержать до 15 цифр.')
            return render(request, 'shop/register.html', get_base_context(request))

        activation_payload = {
            'email': email,
            'first_name': first_name,
            'last_name': last_name,
            'phone_number': normalized_phone,
            'password_hash': make_password(password),
        }
        token = dumps(activation_payload, salt=REGISTER_ACTIVATION_SALT)
        activation_link = request.build_absolute_uri(f'/activate/{token}/')
        email_body = (
            'Здравствуйте!\n\n'
            'Для подтверждения регистрации перейдите по ссылке:\n'
            f'{activation_link}\n\n'
            'Ссылка действительна 1 час.\n'
            'Если вы не регистрировались, просто проигнорируйте это письмо.'
        )

        try:
            send_mail(
                subject='Подтверждение регистрации',
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
        except Exception:
            messages.error(request, 'Не удалось отправить письмо. Проверьте настройки почты.')
            return render(request, 'shop/register.html', get_base_context(request))

        messages.success(request, 'Ссылка для подтверждения регистрации отправлена на вашу почту.')
        return redirect('login')

    return render(request, 'shop/register.html', get_base_context(request))


def login(request):
    next_url = request.POST.get('next') or request.GET.get('next') or ''

    def render_login_page():
        context = get_base_context(request)
        context['next_url'] = next_url
        return render(request, 'shop/login.html', context)

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        password = request.POST.get('password', '')

        if not email or not password:
            messages.error(request, 'Введите email и пароль.')
            return render_login_page()

        user = User.objects.filter(email=email).first()
        if not user:
            messages.error(request, 'Пользователь с такой почтой не найден.')
            return render_login_page()

        valid_password = False
        # Обратная совместимость: если пароль хранится в открытом виде,
        # проверяем его и сразу обновляем на хеш.
        if check_password(password, user.password):
            valid_password = True
        elif password == user.password:
            user.password = make_password(password)
            user.save(update_fields=['password'])
            valid_password = True

        if not valid_password:
            messages.error(request, 'Неверный пароль.')
            return render_login_page()

        request.session['user_id'] = user.id_user
        messages.success(request, f'Вы вошли в аккаунт, {user.first_name}.')
        return redirect(next_url or 'index')

    return render_login_page()


def restore(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        if not email:
            messages.error(request, 'Введите email.')
            return render(request, 'shop/restore.html', get_base_context(request))

        user = User.objects.filter(email=email).first()
        if not user:
            messages.error(request, 'Пользователь с такой почтой не найден.')
            return render(request, 'shop/restore.html', get_base_context(request))

        signer = TimestampSigner()
        token = signer.sign(str(user.id_user))
        reset_link = request.build_absolute_uri(f'/reset-password/{token}/')
        email_body = (
            'Здравствуйте!\n\n'
            'Для сброса пароля перейдите по ссылке:\n'
            f'{reset_link}\n\n'
            'Ссылка действительна 1 час.'
        )
        try:
            send_mail(
                subject='Сброс пароля',
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
        except Exception:
            messages.error(request, 'Не удалось отправить письмо. Попробуйте позже.')
            return render(request, 'shop/restore.html', get_base_context(request))

        messages.success(request, 'Ссылка на сброс отправлена на вашу почту.')
        return redirect('login')

    return render(request, 'shop/restore.html', get_base_context(request))


def activate_account(request, token):
    # Старую сессию завершаем: иначе после перехода по ссылке активации
    # пользователь остаётся «вошедшим» под предыдущим аккаунтом.
    request.session.pop('user_id', None)

    try:
        payload = loads(token, max_age=REGISTER_LINK_TTL_SECONDS, salt=REGISTER_ACTIVATION_SALT)
    except SignatureExpired:
        messages.error(request, 'Ссылка для подтверждения устарела. Зарегистрируйтесь снова.')
        return redirect('register')
    except BadSignature:
        messages.error(request, 'Ссылка для подтверждения недействительна.')
        return redirect('register')

    email = payload.get('email', '').strip().lower()
    if not email:
        messages.error(request, 'Некорректные данные подтверждения. Зарегистрируйтесь снова.')
        return redirect('register')

    if User.objects.filter(email=email).exists():
        messages.info(request, 'Пользователь с такой почтой уже существует. Войдите в аккаунт.')
        return redirect('login')

    try:
        User.objects.create(
            email=email,
            password=payload.get('password_hash', ''),
            first_name=payload.get('first_name', '').strip(),
            last_name=payload.get('last_name', '').strip(),
            phone_number=normalize_phone_number(payload.get('phone_number')) or None,
        )
    except IntegrityError:
        messages.error(
            request,
            'Не удалось завершить регистрацию: проверьте имя, фамилию и телефон (только цифры).'
        )
        return redirect('register')

    messages.success(request, 'Регистрация подтверждена. Теперь вы можете войти.')
    return redirect('login')


def reset_password_confirm(request, token):
    signer = TimestampSigner()
    try:
        user_id = signer.unsign(token, max_age=60 * 60)
    except (BadSignature, SignatureExpired):
        messages.error(request, 'Ссылка для сброса недействительна или устарела.')
        return redirect('restore')

    user = User.objects.filter(id_user=user_id).first()
    if not user:
        messages.error(request, 'Пользователь не найден.')
        return redirect('restore')

    if request.method == 'POST':
        password = request.POST.get('password', '')
        confirm_password = request.POST.get('confirm_password', '')
        if not password or not confirm_password:
            messages.error(request, 'Заполните оба поля пароля.')
        elif password != confirm_password:
            messages.error(request, 'Пароли не совпадают.')
        else:
            user.password = make_password(password)
            user.save(update_fields=['password'])
            messages.success(request, 'Пароль успешно изменен. Войдите с новым паролем.')
            return redirect('login')

    return render(request, 'shop/reset_password_confirm.html', get_base_context(request))


def profile(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Сначала войдите в аккаунт.')
        return redirect_to_login_with_next(request)

    orders = list(
        Order.objects.filter(id_user=user)
        .select_related('id_status', 'id_address__id_city')
        .prefetch_related('orderitem_set__id_product')
        .order_by('-order_date')
    )

    for order in orders:
        order.items_total = 0
        order_rows = []
        for item in order.orderitem_set.all():
            try:
                product = item.id_product
                product_name = product.product_name
                product_price = product.price
            except ObjectDoesNotExist:
                # Показываем старые заказы, даже если товар был удалён.
                product = None
                product_name = 'Товар удален'
                product_price = 0

            line_total = item.quantity * product_price
            order.items_total += line_total
            order_rows.append({
                'item': item,
                'product': product,
                'product_name': product_name,
                'product_price': product_price,
                'line_total': line_total,
            })
        order.rows = order_rows

    context = get_base_context(request)
    context['orders'] = orders
    return render(request, 'shop/profile.html', context)


@require_POST
def profile_update_field(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Сначала войдите в аккаунт.')
        return redirect_to_login_with_next(request)

    field = request.POST.get('field', '').strip()
    value = request.POST.get('value', '').strip()

    if field == 'first_name':
        if not value:
            messages.error(request, 'Имя не может быть пустым.')
        else:
            user.first_name = value
            user.save(update_fields=['first_name'])
            messages.success(request, 'Имя обновлено.')
        return redirect('profile')

    if field == 'last_name':
        if not value:
            messages.error(request, 'Фамилия не может быть пустой.')
        else:
            user.last_name = value
            user.save(update_fields=['last_name'])
            messages.success(request, 'Фамилия обновлена.')
        return redirect('profile')

    if field == 'phone_number':
        normalized_phone = normalize_phone_number(value)
        if value and not normalized_phone:
            messages.error(request, 'Телефон должен содержать до 15 цифр.')
        else:
            user.phone_number = normalized_phone
            user.save(update_fields=['phone_number'])
            messages.success(request, 'Телефон обновлен.')
        return redirect('profile')

    if field == 'email':
        new_email = value.lower()
        if not new_email:
            messages.error(request, 'Email не может быть пустым.')
            return redirect('profile')
        if new_email == user.email:
            messages.info(request, 'Вы ввели текущий email. Изменений нет.')
            return redirect('profile')
        if User.objects.filter(email=new_email).exclude(id_user=user.id_user).exists():
            messages.error(request, 'Пользователь с таким email уже существует.')
            return redirect('profile')

        token = dumps(
            {
                'user_id': user.id_user,
                'new_email': new_email,
            },
            salt=PROFILE_EMAIL_CHANGE_SALT
        )
        confirm_link = request.build_absolute_uri(
            reverse('confirm_profile_email_change', kwargs={'token': token})
        )
        email_body = (
            'Здравствуйте!\n\n'
            'Вы запросили изменение email в профиле.\n'
            'Подтвердите новый адрес по ссылке:\n'
            f'{confirm_link}\n\n'
            'Ссылка действительна 1 час.\n'
            'Если это были не вы, просто проигнорируйте письмо.'
        )
        try:
            send_mail(
                subject='Подтверждение смены email',
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[new_email],
                fail_silently=False,
            )
            messages.success(request, 'Ссылка для подтверждения нового email отправлена.')
        except Exception:
            messages.error(request, 'Не удалось отправить письмо подтверждения. Попробуйте позже.')
        return redirect('profile')

    messages.error(request, 'Неизвестное поле для обновления.')
    return redirect('profile')


def confirm_profile_email_change(request, token):
    try:
        payload = loads(
            token,
            max_age=PROFILE_EMAIL_CHANGE_LINK_TTL_SECONDS,
            salt=PROFILE_EMAIL_CHANGE_SALT
        )
    except SignatureExpired:
        messages.error(request, 'Ссылка подтверждения устарела. Запросите изменение email снова.')
        return redirect('profile')
    except BadSignature:
        messages.error(request, 'Ссылка подтверждения недействительна.')
        return redirect('profile')

    user_id = payload.get('user_id')
    new_email = str(payload.get('new_email', '')).strip().lower()
    user = User.objects.filter(id_user=user_id).first()
    if not user or not new_email:
        messages.error(request, 'Некорректные данные подтверждения email.')
        return redirect('profile')

    if User.objects.filter(email=new_email).exclude(id_user=user.id_user).exists():
        messages.error(request, 'Этот email уже используется другим пользователем.')
        return redirect('profile')

    user.email = new_email
    user.save(update_fields=['email'])
    messages.success(request, 'Email успешно подтвержден и обновлен.')
    return redirect('profile')


def manager_dashboard(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Для доступа в панель администратора войдите в аккаунт.')
        return redirect_to_login_with_next(request)
    if not can_manage_store(user):
        messages.error(request, 'Доступ разрешен только администраторам.')
        return redirect('index')

    tab = request.GET.get('tab', 'orders')
    statuses = OrderStatus.objects.order_by('id_status')
    categories = Category.objects.order_by('category_name')
    parent_category_ids = set(
        Category.objects.filter(id_parent_category__isnull=False)
        .values_list('id_parent_category_id', flat=True)
    )
    leaf_categories = categories.exclude(id_category__in=parent_category_ids)
    categories_without_products = categories.exclude(
        id_category__in=Product.objects.values_list('id_category_id', flat=True).distinct()
    )
    admin_categories = categories_without_products.filter(id_category__in=parent_category_ids)

    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        if action == 'update_order_status':
            order_id = request.POST.get('order_id', '').strip()
            status_id = request.POST.get('status_id', '').strip()
            order = Order.objects.filter(id_order=order_id).first()
            status = statuses.filter(id_status=status_id).first()
            if not order or not status:
                messages.error(request, 'Не удалось обновить статус заказа.')
            else:
                order.id_status = status
                order.save(update_fields=['id_status'])
                messages.success(request, 'Статус заказа обновлен.')
            return redirect(f"{reverse('manager_dashboard')}?tab=orders")

        if action == 'create_product':
            name = request.POST.get('product_name', '').strip()
            price_raw = request.POST.get('price', '0').strip()
            stock_raw = request.POST.get('stock_quantity', '0').strip()
            category_id = request.POST.get('category_id', '').strip()
            description = request.POST.get('description', '').strip()
            image_url = request.POST.get('image_url', '').strip()

            category = leaf_categories.filter(id_category=category_id).first()
            try:
                price = int(price_raw)
                stock_quantity = int(stock_raw)
            except (TypeError, ValueError):
                price = -1
                stock_quantity = -1

            if not name or not category or price < 0 or stock_quantity < 0:
                messages.error(request, 'Проверьте данные нового товара: название, листовая категория, цена и остаток.')
            else:
                product = Product.objects.create(
                    product_name=name,
                    description=description or None,
                    price=price,
                    stock_quantity=stock_quantity,
                    id_category=category,
                    image_url=image_url or None,
                )
                mappings = list(
                    CategoryAttributeMap.objects.select_related('id_attribute')
                    .filter(id_category=category)
                    .order_by('step_order', 'id_category_attribute_map')
                )
                for mapping in mappings:
                    attribute = mapping.id_attribute
                    if not attribute or not attribute.is_active:
                        continue
                    field_name = f'attr_{attribute.id_attribute}'
                    raw_value = request.POST.get(field_name, '').strip()
                    if attribute.value_type == AttributeDefinition.VALUE_TYPE_ENUM:
                        if raw_value == '__new__':
                            new_option_value = request.POST.get(f'attr_new_{attribute.id_attribute}', '').strip()
                            if not new_option_value:
                                continue
                            option, _ = AttributeOption.objects.get_or_create(
                                id_attribute=attribute,
                                option_value=new_option_value,
                                defaults={'sort_order': 100, 'is_active': True},
                            )
                            ProductAttributeValue.objects.update_or_create(
                                id_product=product,
                                id_attribute=attribute,
                                defaults={'id_option': option}
                            )
                        elif raw_value.isdigit():
                            option = AttributeOption.objects.filter(
                                id_option=int(raw_value),
                                id_attribute=attribute,
                            ).first()
                            if option:
                                ProductAttributeValue.objects.update_or_create(
                                    id_product=product,
                                    id_attribute=attribute,
                                    defaults={'id_option': option}
                                )
                    elif attribute.value_type == AttributeDefinition.VALUE_TYPE_NUMBER:
                        if not raw_value:
                            continue
                        normalized = raw_value.replace(',', '.')
                        try:
                            number_value = float(normalized)
                        except (TypeError, ValueError):
                            continue
                        ProductAttributeValue.objects.update_or_create(
                            id_product=product,
                            id_attribute=attribute,
                            defaults={'value_number': number_value}
                        )
                    elif attribute.value_type == AttributeDefinition.VALUE_TYPE_BOOLEAN:
                        if raw_value not in ('0', '1'):
                            continue
                        ProductAttributeValue.objects.update_or_create(
                            id_product=product,
                            id_attribute=attribute,
                            defaults={'value_boolean': raw_value == '1'}
                        )
                    else:
                        if not raw_value:
                            continue
                        ProductAttributeValue.objects.update_or_create(
                            id_product=product,
                            id_attribute=attribute,
                            defaults={'value_string': raw_value}
                        )
                messages.success(request, 'Товар добавлен.')
            return redirect(f"{reverse('manager_dashboard')}?tab=catalog")

        if action == 'update_product':
            product_id = request.POST.get('product_id', '').strip()
            product = Product.objects.filter(id_product=product_id).first()
            if not product:
                messages.error(request, 'Товар не найден.')
                return redirect(f"{reverse('manager_dashboard')}?tab=catalog")

            name = request.POST.get('product_name', '').strip()
            description = request.POST.get('description', '').strip()
            image_url = request.POST.get('image_url', '').strip()
            category_id = request.POST.get('category_id', '').strip()
            category = categories.filter(id_category=category_id).first()
            try:
                price = int(request.POST.get('price', '0').strip())
                stock_quantity = int(request.POST.get('stock_quantity', '0').strip())
            except (TypeError, ValueError):
                price = -1
                stock_quantity = -1

            if not name or not category or price < 0 or stock_quantity < 0:
                messages.error(request, f'Некорректные данные товара #{product_id}.')
            else:
                product.product_name = name
                product.description = description or None
                product.image_url = image_url or None
                product.id_category = category
                product.price = price
                product.stock_quantity = stock_quantity
                product.save()
                messages.success(request, f'Товар #{product_id} обновлен.')
            return redirect(f"{reverse('manager_dashboard')}?tab=catalog")

        if action == 'import_analytics_xlsx':
            uploaded_file = request.FILES.get('analytics_file')
            if not uploaded_file:
                messages.error(request, 'Выберите XLSX-файл аналитики для импорта.')
                return redirect(f"{reverse('manager_dashboard')}?tab=analytics")
            if not uploaded_file.name.lower().endswith('.xlsx'):
                messages.error(request, 'Поддерживается только формат .xlsx.')
                return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

            try:
                workbook = load_workbook(uploaded_file, data_only=True)
            except Exception:
                messages.error(request, 'Не удалось прочитать файл. Проверьте, что это корректный XLSX.')
                return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

            required_sheets = {'Продажи по дням', 'Категории', 'Топ товаров'}
            if not required_sheets.issubset(set(workbook.sheetnames)):
                messages.error(
                    request,
                    'Файл аналитики не подходит: должны быть листы "Продажи по дням", "Категории" и "Топ товаров".'
                )
                return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

            daily_labels = []
            daily_totals = []
            for row in workbook['Продажи по дням'].iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                day_label = format_day_label(row[0] if len(row) > 0 else None)
                total_value = parse_non_negative_int(row[1] if len(row) > 1 else None)
                if not day_label or total_value is None:
                    continue
                daily_labels.append(day_label)
                daily_totals.append(total_value)

            category_labels = []
            category_totals = []
            for row in workbook['Категории'].iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                label = str(row[0] if len(row) > 0 else '').strip()
                total_value = parse_float_value(row[1] if len(row) > 1 else None)
                if not label or total_value is None:
                    continue
                category_labels.append(label)
                category_totals.append(total_value)

            imported_top_products = []
            for row in workbook['Топ товаров'].iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                product_name = str(row[0] if len(row) > 0 else '').strip()
                qty = parse_non_negative_int(row[1] if len(row) > 1 else None)
                sales = parse_float_value(row[2] if len(row) > 2 else None)
                if not product_name or qty is None or sales is None:
                    continue
                imported_top_products.append({
                    'id_product__product_name': product_name,
                    'total_quantity': qty,
                    'total_sales': sales,
                })

            if not daily_labels and not category_labels and not imported_top_products:
                messages.error(request, 'Не удалось импортировать аналитические данные: файл пуст или поврежден.')
                return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

            request.session['imported_analytics_snapshot'] = {
                'daily_labels': daily_labels,
                'daily_totals': daily_totals,
                'category_labels': category_labels,
                'category_totals': category_totals,
                'top_products': imported_top_products,
            }
            messages.success(request, 'Аналитика успешно импортирована из XLSX и отображается на этой вкладке.')
            return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

        if action == 'clear_imported_analytics':
            request.session.pop('imported_analytics_snapshot', None)
            messages.success(request, 'Импортированная аналитика очищена. Показаны актуальные данные из системы.')
            return redirect(f"{reverse('manager_dashboard')}?tab=analytics")

        if is_admin(user) and action == 'create_category':
            category_name = request.POST.get('category_name', '').strip()
            parent_category_id = request.POST.get('parent_category_id', '').strip()
            parent_category = categories.filter(id_category=parent_category_id).first() if parent_category_id else None
            if not category_name:
                messages.error(request, 'Название категории не может быть пустым.')
            else:
                Category.objects.create(category_name=category_name, id_parent_category=parent_category)
                messages.success(request, 'Категория добавлена.')
            return redirect(f"{reverse('manager_dashboard')}?tab=categories")

        if is_admin(user) and action == 'delete_category':
            category_id = request.POST.get('category_id', '').strip()
            category_to_delete = categories.filter(id_category=category_id).first()
            if not category_to_delete:
                messages.error(request, 'Категория не найдена.')
            elif Product.objects.filter(id_category=category_to_delete).exists():
                messages.error(request, 'Категорию нельзя удалить: в ней есть товары.')
            elif Category.objects.filter(id_parent_category=category_to_delete).exists():
                messages.error(request, 'Категорию нельзя удалить: у нее есть подкатегории.')
            else:
                category_to_delete.delete()
                messages.success(request, 'Категория удалена.')
            return redirect(f"{reverse('manager_dashboard')}?tab=categories")

        if is_admin(user) and action == 'update_category_name':
            category_id = request.POST.get('category_id', '').strip()
            new_name = request.POST.get('category_name', '').strip()
            category_to_update = categories.filter(id_category=category_id).first()
            if not category_to_update:
                messages.error(request, 'Категория не найдена.')
            elif not new_name:
                messages.error(request, 'Название категории не может быть пустым.')
            else:
                category_to_update.category_name = new_name
                category_to_update.save(update_fields=['category_name'])
                messages.success(request, 'Название категории обновлено.')
            return redirect(f"{reverse('manager_dashboard')}?tab=categories")

        if is_admin(user) and action == 'update_user_role':
            target_user_id = request.POST.get('target_user_id', '').strip()
            new_role = request.POST.get('new_role', '').strip()
            target_user = User.objects.filter(id_user=target_user_id).first()
            allowed_roles = {choice[0] for choice in User.ROLE_CHOICES}
            if not target_user or new_role not in allowed_roles:
                messages.error(request, 'Некорректные данные для смены роли пользователя.')
            elif target_user.id_user == user.id_user and new_role != User.ROLE_ADMIN:
                messages.error(request, 'Нельзя снять роль администратора с текущей сессии.')
            else:
                target_user.role = new_role
                target_user.save(update_fields=['role'])
                messages.success(request, 'Роль пользователя обновлена.')
            return redirect(f"{reverse('manager_dashboard')}?tab=users")

        if is_admin(user) and action == 'update_user_profile':
            target_user_id = request.POST.get('target_user_id', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip().lower()
            target_user = User.objects.filter(id_user=target_user_id).first()
            if not target_user:
                messages.error(request, 'Пользователь не найден.')
            elif not first_name or not last_name or not email:
                messages.error(request, 'Имя, фамилия и email обязательны.')
            elif User.objects.filter(email=email).exclude(id_user=target_user.id_user).exists():
                messages.error(request, 'Такой email уже используется.')
            else:
                target_user.first_name = first_name
                target_user.last_name = last_name
                target_user.email = email
                target_user.save(update_fields=['first_name', 'last_name', 'email'])
                messages.success(request, 'Данные пользователя обновлены.')
            return redirect(f"{reverse('manager_dashboard')}?tab=users")

        if is_admin(user) and action == 'delete_user':
            target_user_id = request.POST.get('target_user_id', '').strip()
            target_user = User.objects.filter(id_user=target_user_id).first()
            if not target_user:
                messages.error(request, 'Пользователь не найден.')
            elif target_user.id_user == user.id_user:
                messages.error(request, 'Нельзя удалить текущего администратора.')
            else:
                target_user.delete()
                messages.success(request, 'Пользователь удален.')
            return redirect(f"{reverse('manager_dashboard')}?tab=users")

        if is_admin(user) and action == 'delete_review':
            review_id = request.POST.get('review_id', '').strip()
            review = Review.objects.filter(review_id=review_id).first()
            if not review:
                messages.error(request, 'Отзыв не найден.')
            else:
                review.delete()
                messages.success(request, 'Отзыв удален.')
            return redirect(f"{reverse('manager_dashboard')}?tab=reviews")

    orders = list(
        Order.objects.select_related(
        'id_user',
        'id_address__id_city',
        'id_status',
    ).prefetch_related(
        'orderitem_set__id_product',
    ).order_by('-order_date')
    )
    active_orders = [
        order for order in orders
        if (order.id_status and (order.id_status.status_name or '').strip().lower() != 'доставлен')
    ]
    delivered_orders = [
        order for order in orders
        if (order.id_status and (order.id_status.status_name or '').strip().lower() == 'доставлен')
    ]

    products = Product.objects.select_related('id_category').order_by('-id_product')

    start_date = timezone.localdate() - timedelta(days=29)
    sales_by_day = (
        Order.objects.filter(order_date__date__gte=start_date)
        .annotate(day=TruncDate('order_date'))
        .values('day')
        .annotate(total=Sum('total_cost'))
        .order_by('day')
    )
    totals_by_day = {
        entry['day']: int(entry['total'] or 0)
        for entry in sales_by_day
        if entry['day']
    }
    all_days = [start_date + timedelta(days=offset) for offset in range(30)]
    daily_labels = [day.strftime('%d.%m') for day in all_days]
    daily_totals = [totals_by_day.get(day, 0) for day in all_days]

    revenue_expr = ExpressionWrapper(
        F('quantity') * F('id_product__price'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    category_sales = (
        OrderItem.objects.values('id_product__id_category')
        .annotate(total_sales=Sum(revenue_expr))
    )
    categories_by_id = {
        category.id_category: category
        for category in Category.objects.only('id_category', 'category_name', 'id_parent_category')
    }

    def get_root_category(category_id):
        current = categories_by_id.get(category_id)
        visited = set()
        while current and current.id_parent_category_id and current.id_category not in visited:
            visited.add(current.id_category)
            current = categories_by_id.get(current.id_parent_category_id)
        return current

    root_category_totals = defaultdict(float)
    for entry in category_sales:
        category_id = entry['id_product__id_category']
        total_sales = float(entry['total_sales'] or 0)
        root_category = get_root_category(category_id)
        root_label = root_category.category_name if root_category else 'Без категории'
        root_category_totals[root_label] += total_sales

    top_root_categories = sorted(root_category_totals.items(), key=lambda item: item[1], reverse=True)[:5]
    category_labels = [label for label, _ in top_root_categories]
    category_totals = [total for _, total in top_root_categories]

    top_products = (
        OrderItem.objects.values('id_product__product_name')
        .annotate(total_quantity=Sum('quantity'), total_sales=Sum(revenue_expr))
        .order_by('-total_quantity')[:10]
    )

    imported_analytics = request.session.get('imported_analytics_snapshot')
    has_imported_analytics = bool(imported_analytics)
    if imported_analytics:
        daily_labels = imported_analytics.get('daily_labels', daily_labels)
        daily_totals = imported_analytics.get('daily_totals', daily_totals)
        category_labels = imported_analytics.get('category_labels', category_labels)
        category_totals = imported_analytics.get('category_totals', category_totals)
        top_products = imported_analytics.get('top_products', list(top_products))

    attribute_mappings = (
        CategoryAttributeMap.objects.select_related('id_category', 'id_attribute')
        .order_by('id_category_id', 'step_order', 'id_category_attribute_map')
    )
    attribute_options_map = defaultdict(list)
    for option in AttributeOption.objects.filter(is_active=True).order_by('sort_order', 'option_value'):
        attribute_options_map[option.id_attribute_id].append({
            'id': option.id_option,
            'value': option.option_value,
        })
    category_attribute_config = defaultdict(list)
    for mapping in attribute_mappings:
        attribute = mapping.id_attribute
        if not attribute or not attribute.is_active:
            continue
        if mapping.id_category_id not in parent_category_ids:
            category_attribute_config[mapping.id_category_id].append({
                'id': attribute.id_attribute,
                'name': attribute.attribute_name,
                'code': attribute.attribute_code,
                'value_type': attribute.value_type,
                'unit': attribute.unit or '',
                'options': attribute_options_map.get(attribute.id_attribute, []),
            })

    admin_users = User.objects.order_by('-id_user')
    admin_reviews = Review.objects.select_related('id_user', 'id_product').order_by('-review_date')[:100]

    return render(request, 'shop/manager_dashboard.html', {
        'current_user': user,
        'cart_item_count': CartService.get_cart_item_count(user),
        'is_admin': is_admin(user),
        'tab': tab,
        'active_orders': active_orders,
        'delivered_orders': delivered_orders,
        'statuses': statuses,
        'products': products,
        'categories': categories,
        'leaf_categories': leaf_categories,
        'admin_categories': admin_categories,
        'admin_users': admin_users,
        'admin_reviews': admin_reviews,
        'category_attribute_config': dict(category_attribute_config),
        'daily_labels': daily_labels,
        'daily_totals': daily_totals,
        'category_labels': category_labels,
        'category_totals': category_totals,
        'top_products': top_products,
        'has_imported_analytics': has_imported_analytics,
    })


def manager_export_analytics_xlsx(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Для доступа в панель администратора войдите в аккаунт.')
        return redirect_to_login_with_next(request)
    if not can_manage_store(user):
        messages.error(request, 'Доступ разрешен только администраторам.')
        return redirect('index')

    start_date = timezone.localdate() - timedelta(days=29)
    sales_by_day = (
        Order.objects.filter(order_date__date__gte=start_date)
        .annotate(day=TruncDate('order_date'))
        .values('day')
        .annotate(total=Sum('total_cost'))
        .order_by('day')
    )
    totals_by_day = {
        entry['day']: int(entry['total'] or 0)
        for entry in sales_by_day
        if entry['day']
    }
    all_days = [start_date + timedelta(days=offset) for offset in range(30)]

    revenue_expr = ExpressionWrapper(
        F('quantity') * F('id_product__price'),
        output_field=DecimalField(max_digits=14, decimal_places=2),
    )
    category_sales = (
        OrderItem.objects.values('id_product__id_category')
        .annotate(total_sales=Sum(revenue_expr))
    )
    categories_by_id = {
        category.id_category: category
        for category in Category.objects.only('id_category', 'category_name', 'id_parent_category')
    }

    def get_root_category(category_id):
        current = categories_by_id.get(category_id)
        visited = set()
        while current and current.id_parent_category_id and current.id_category not in visited:
            visited.add(current.id_category)
            current = categories_by_id.get(current.id_parent_category_id)
        return current

    root_category_totals = defaultdict(float)
    for entry in category_sales:
        category_id = entry['id_product__id_category']
        total_sales = float(entry['total_sales'] or 0)
        root_category = get_root_category(category_id)
        root_label = root_category.category_name if root_category else 'Без категории'
        root_category_totals[root_label] += total_sales

    top_root_categories = sorted(root_category_totals.items(), key=lambda item: item[1], reverse=True)[:5]
    top_products = list(
        OrderItem.objects.values('id_product__product_name')
        .annotate(total_quantity=Sum('quantity'), total_sales=Sum(revenue_expr))
        .order_by('-total_quantity')[:10]
    )

    workbook = Workbook()
    sheet_daily = workbook.active
    sheet_daily.title = 'Продажи по дням'
    sheet_daily.append(['Дата', 'Выручка (₽)'])
    for day in all_days:
        sheet_daily.append([day.strftime('%d.%m.%Y'), totals_by_day.get(day, 0)])

    sheet_categories = workbook.create_sheet(title='Категории')
    sheet_categories.append(['Категория', 'Сумма продаж (₽)'])
    for label, total in top_root_categories:
        sheet_categories.append([label, round(total, 2)])

    sheet_products = workbook.create_sheet(title='Топ товаров')
    sheet_products.append(['Товар', 'Количество', 'Сумма продаж (₽)'])
    for item in top_products:
        sheet_products.append([
            item.get('id_product__product_name') or 'Без названия',
            int(item.get('total_quantity') or 0),
            float(item.get('total_sales') or 0),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'analytics_{timezone.localdate().strftime("%Y-%m-%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def manager_analytics_template_xlsx(request):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Для доступа в панель администратора войдите в аккаунт.')
        return redirect_to_login_with_next(request)
    if not can_manage_store(user):
        messages.error(request, 'Доступ разрешен только администраторам.')
        return redirect('index')

    workbook = Workbook()

    sheet_daily = workbook.active
    sheet_daily.title = 'Продажи по дням'
    sheet_daily.append(['Дата', 'Выручка (₽)'])
    sheet_daily.append(['01.01.2026', 120000])
    sheet_daily.append(['02.01.2026', 98000])

    sheet_categories = workbook.create_sheet(title='Категории')
    sheet_categories.append(['Категория', 'Сумма продаж (₽)'])
    sheet_categories.append(['Смартфоны и гаджеты', 450000])
    sheet_categories.append(['Ноутбуки', 380000])

    sheet_products = workbook.create_sheet(title='Топ товаров')
    sheet_products.append(['Товар', 'Количество', 'Сумма продаж (₽)'])
    sheet_products.append(['Apple iPhone 17 Pro 256GB', 15, 2000000])
    sheet_products.append(['HONOR MagicBook X16', 10, 520000])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="analytics_import_template.xlsx"'
    workbook.save(response)
    return response


@require_POST
def manager_update_product(request, product_id):
    user = get_logged_in_user(request)
    if not user:
        messages.info(request, 'Для редактирования товара войдите в аккаунт.')
        return redirect_to_login_with_next(request)
    if not can_manage_store(user):
        messages.error(request, 'Редактировать товары может только администратор.')
        return redirect('index')

    product = Product.objects.filter(id_product=product_id).first()
    if not product:
        messages.error(request, 'Товар не найден.')
        return redirect('product_detail', product_id=product_id)

    action = request.POST.get('action', 'update').strip()
    if action == 'delete':
        product.delete()
        messages.success(request, 'Товар удален.')
        return redirect(f"{reverse('manager_dashboard')}?tab=catalog")

    name = request.POST.get('product_name', '').strip()
    description = request.POST.get('description', '').strip()
    image_url = request.POST.get('image_url', '').strip()
    category_id = request.POST.get('category_id', '').strip()
    parent_category_ids = set(
        Category.objects.filter(id_parent_category__isnull=False)
        .values_list('id_parent_category_id', flat=True)
    )
    category = Category.objects.exclude(id_category__in=parent_category_ids).filter(id_category=category_id).first()
    try:
        price = int(request.POST.get('price', '0').strip())
        stock_quantity = int(request.POST.get('stock_quantity', '0').strip())
    except (TypeError, ValueError):
        price = -1
        stock_quantity = -1

    if not name or not category or price < 0 or stock_quantity < 0:
        messages.error(request, 'Некорректные данные товара. Проверьте поля и попробуйте снова.')
        return redirect('product_detail', product_id=product.id_product)

    product.product_name = name
    product.description = description or None
    product.image_url = image_url or None
    product.id_category = category
    product.price = price
    product.stock_quantity = stock_quantity
    product.save()

    mappings = list(
        CategoryAttributeMap.objects.select_related('id_attribute')
        .filter(id_category=category)
        .order_by('step_order', 'id_category_attribute_map')
    )
    allowed_attribute_ids = set()
    for mapping in mappings:
        attribute = mapping.id_attribute
        if not attribute or not attribute.is_active:
            continue
        allowed_attribute_ids.add(attribute.id_attribute)
        field_name = f'attr_{attribute.id_attribute}'
        raw_value = request.POST.get(field_name, '').strip()
        if attribute.value_type == AttributeDefinition.VALUE_TYPE_ENUM:
            if raw_value == '__new__':
                new_option_value = request.POST.get(f'attr_new_{attribute.id_attribute}', '').strip()
                if not new_option_value:
                    ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
                    continue
                option, _ = AttributeOption.objects.get_or_create(
                    id_attribute=attribute,
                    option_value=new_option_value,
                    defaults={'sort_order': 100, 'is_active': True},
                )
                ProductAttributeValue.objects.update_or_create(
                    id_product=product,
                    id_attribute=attribute,
                    defaults={'id_option': option, 'value_number': None, 'value_string': None, 'value_boolean': None}
                )
            elif raw_value.isdigit():
                option = AttributeOption.objects.filter(id_option=int(raw_value), id_attribute=attribute).first()
                if option:
                    ProductAttributeValue.objects.update_or_create(
                        id_product=product,
                        id_attribute=attribute,
                        defaults={'id_option': option, 'value_number': None, 'value_string': None, 'value_boolean': None}
                    )
                else:
                    ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
            else:
                ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
        elif attribute.value_type == AttributeDefinition.VALUE_TYPE_NUMBER:
            if not raw_value:
                ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
                continue
            normalized = raw_value.replace(',', '.')
            try:
                number_value = float(normalized)
            except (TypeError, ValueError):
                continue
            ProductAttributeValue.objects.update_or_create(
                id_product=product,
                id_attribute=attribute,
                defaults={'value_number': number_value, 'id_option': None, 'value_string': None, 'value_boolean': None}
            )
        elif attribute.value_type == AttributeDefinition.VALUE_TYPE_BOOLEAN:
            if raw_value not in ('0', '1'):
                ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
                continue
            ProductAttributeValue.objects.update_or_create(
                id_product=product,
                id_attribute=attribute,
                defaults={'value_boolean': raw_value == '1', 'id_option': None, 'value_number': None, 'value_string': None}
            )
        else:
            if not raw_value:
                ProductAttributeValue.objects.filter(id_product=product, id_attribute=attribute).delete()
                continue
            ProductAttributeValue.objects.update_or_create(
                id_product=product,
                id_attribute=attribute,
                defaults={'value_string': raw_value, 'id_option': None, 'value_number': None, 'value_boolean': None}
            )

    ProductAttributeValue.objects.filter(id_product=product).exclude(id_attribute_id__in=allowed_attribute_ids).delete()

    messages.success(request, 'Товар успешно обновлен.')
    return redirect('product_detail', product_id=product.id_product)


def logout(request):
    request.session.pop('user_id', None)
    messages.success(request, 'Вы вышли из аккаунта.')
    return redirect('index')

def categories(request):
    categories_list = CategoryService.get_root_categories_ordered()
    categories_list = CategoryService.enrich_categories_with_images(categories_list)
    
    user = UserService.get_user_from_session(request)
    cart_item_count = CartService.get_cart_item_count(user)
    
    return render(request, 'shop/categories.html', {
        'categories': categories_list,
        'cart_item_count': cart_item_count,
        'current_user': user,
    })

def product_list(request):
    search_query = request.GET.get('q', '').strip()
    search_products = ProductService.search_products(search_query)
    bounds = search_products.aggregate(min_price=Min('price'), max_price=Max('price'))
    bounds_min_price = bounds.get('min_price') if bounds.get('min_price') is not None else 0
    bounds_max_price = bounds.get('max_price') if bounds.get('max_price') is not None else 0
    products = search_products
    min_price_raw = request.GET.get('min_price', '').strip()
    max_price_raw = request.GET.get('max_price', '').strip()

    min_price = None
    max_price = None
    try:
        if min_price_raw:
            min_price = max(0, int(min_price_raw))
    except (TypeError, ValueError):
        min_price = None

    try:
        if max_price_raw:
            max_price = max(0, int(max_price_raw))
    except (TypeError, ValueError):
        max_price = None

    if min_price is not None and max_price is not None and min_price > max_price:
        min_price, max_price = max_price, min_price
        min_price_raw, max_price_raw = str(min_price), str(max_price)

    if min_price is not None:
        products = products.filter(price__gte=min_price)
    if max_price is not None:
        products = products.filter(price__lte=max_price)

    categories_list = CategoryService.get_root_categories()
    
    user = UserService.get_user_from_session(request)
    cart_item_count = CartService.get_cart_item_count(user)
    
    return render(request, 'shop/product_list.html', {
        'products': products,
        'categories': categories_list,
        'cart_item_count': cart_item_count,
        'current_user': user,
        'search_query': search_query,
        'min_price': min_price_raw or str(bounds_min_price),
        'max_price': max_price_raw or str(bounds_max_price),
        'search_min_price': bounds_min_price,
        'search_max_price': bounds_max_price,
    })

def category_detail(request, category_id):
    category = get_object_or_404(Category, id_category=category_id)
    category.card_image = CategoryService.get_category_image(category.category_name)
    root_category = category
    while root_category and root_category.id_parent_category:
        root_category = root_category.id_parent_category
    root_category_name = root_category.category_name if root_category else ''
    
    children = CategoryService.get_subcategories(category)
    children = CategoryService.enrich_categories_with_images(children)
    
    products = []
    category_filters = []
    min_price_raw = request.GET.get('min_price', '').strip()
    max_price_raw = request.GET.get('max_price', '').strip()
    min_price_value = None
    max_price_value = None
    selected_attribute_filters = {}
    price_bounds_min = 0
    price_bounds_max = 0

    def parse_int(raw_value):
        try:
            return max(0, int(raw_value))
        except (TypeError, ValueError):
            return None

    def parse_decimal(raw_value):
        if raw_value is None or raw_value == '':
            return None
        normalized = str(raw_value).replace(',', '.').strip()
        try:
            return float(normalized)
        except (TypeError, ValueError):
            return None

    if not children:
        products = ProductService.get_products_by_category(category)
        base_products = products
        price_bounds = base_products.aggregate(min_price=Min('price'), max_price=Max('price'))
        price_bounds_min = int(price_bounds.get('min_price') or 0)
        price_bounds_max = int(price_bounds.get('max_price') or 0)

        min_price_value = parse_int(min_price_raw)
        max_price_value = parse_int(max_price_raw)
        if min_price_value is not None and max_price_value is not None and min_price_value > max_price_value:
            min_price_value, max_price_value = max_price_value, min_price_value
            min_price_raw, max_price_raw = str(min_price_value), str(max_price_value)

        if min_price_value is not None:
            products = products.filter(price__gte=min_price_value)
        if max_price_value is not None:
            products = products.filter(price__lte=max_price_value)

        mappings = list(
            CategoryAttributeMap.objects
            .select_related('id_attribute')
            .filter(id_category=category)
            .order_by('step_order', 'id_category_attribute_map')
        )
        base_product_ids = base_products.values('id_product')
        products_qs = products
        for mapping in mappings:
            attribute = mapping.id_attribute
            if not attribute or not attribute.is_active:
                continue
            if attribute.attribute_code == 'brand' and category.category_name in ('Apple iPhone', 'Samsung Galaxy'):
                continue
            is_filterable = (
                mapping.is_filterable_override
                if mapping.is_filterable_override is not None
                else attribute.is_filterable
            )
            if not is_filterable:
                continue

            filter_item = {
                'attribute_id': attribute.id_attribute,
                'attribute_name': attribute.attribute_name,
                'attribute_code': attribute.attribute_code,
                'value_type': attribute.value_type,
                'unit': attribute.unit or '',
            }

            if attribute.value_type == AttributeDefinition.VALUE_TYPE_ENUM:
                available_option_ids = (
                    ProductAttributeValue.objects
                    .filter(id_product_id__in=base_product_ids, id_attribute=attribute, id_option__isnull=False)
                    .values_list('id_option_id', flat=True)
                    .distinct()
                )
                options = list(
                    AttributeOption.objects
                    .filter(id_option__in=available_option_ids, id_attribute=attribute, is_active=True)
                    .order_by('sort_order', 'option_value')
                )
                if not options:
                    continue
                param_key = f'attr_{attribute.id_attribute}'
                selected_values = [value for value in request.GET.getlist(param_key) if value]
                selected_option_ids = [int(value) for value in selected_values if value.isdigit()]
                if selected_option_ids:
                    matching_ids = ProductAttributeValue.objects.filter(
                        id_product_id__in=products_qs.values('id_product'),
                        id_attribute=attribute,
                        id_option_id__in=selected_option_ids,
                    ).values('id_product')
                    products_qs = products_qs.filter(id_product__in=matching_ids)
                selected_set = set(str(value) for value in selected_option_ids)
                filter_item['param_key'] = param_key
                filter_item['options'] = [
                    {
                        'id': option.id_option,
                        'value': option.option_value,
                        'selected': str(option.id_option) in selected_set,
                    }
                    for option in options
                ]
                selected_attribute_filters[param_key] = selected_values

            elif attribute.value_type == AttributeDefinition.VALUE_TYPE_NUMBER:
                use_discrete_number = attribute.attribute_code == 'memory_gb' or (
                    attribute.attribute_code == 'diagonal_inch' and root_category_name == 'Смартфоны и гаджеты'
                ) or (
                    attribute.attribute_code == 'diagonal_inch' and root_category_name in ('Ноутбуки', 'ТВ и консоли')
                ) or (
                    attribute.attribute_code == 'ram_gb' and root_category_name == 'Ноутбуки'
                )
                if use_discrete_number:
                    value_rows = list(
                        ProductAttributeValue.objects.filter(
                            id_product_id__in=base_product_ids,
                            id_attribute=attribute,
                            value_number__isnull=False
                        ).values_list('value_number', flat=True).distinct()
                    )
                    option_values = sorted({
                        int(value) if float(value).is_integer() else float(value)
                        for value in value_rows
                    })
                    if not option_values:
                        continue
                    param_key = f'attr_{attribute.id_attribute}'
                    selected_values = [value for value in request.GET.getlist(param_key) if value]
                    selected_numbers = []
                    for raw in selected_values:
                        parsed = parse_decimal(raw)
                        if parsed is not None:
                            selected_numbers.append(parsed)
                    if selected_numbers:
                        matching_ids = ProductAttributeValue.objects.filter(
                            id_product_id__in=products_qs.values('id_product'),
                            id_attribute=attribute,
                            value_number__in=selected_numbers,
                        ).values('id_product')
                        products_qs = products_qs.filter(id_product__in=matching_ids)

                    selected_set = set(str(value) for value in selected_values)
                    filter_item['ui_type'] = 'discrete_number'
                    filter_item['param_key'] = param_key
                    filter_item['options'] = [
                        {
                            'value': value,
                            'selected': str(value) in selected_set,
                        }
                        for value in option_values
                    ]
                    selected_attribute_filters[param_key] = selected_values
                    category_filters.append(filter_item)
                    continue

                number_bounds = ProductAttributeValue.objects.filter(
                    id_product_id__in=base_product_ids,
                    id_attribute=attribute,
                    value_number__isnull=False
                ).aggregate(min_value=Min('value_number'), max_value=Max('value_number'))
                if number_bounds['min_value'] is None:
                    continue
                min_key = f'attr_{attribute.id_attribute}_min'
                max_key = f'attr_{attribute.id_attribute}_max'
                selected_min_raw = request.GET.get(min_key, '').strip()
                selected_max_raw = request.GET.get(max_key, '').strip()
                selected_min = parse_decimal(selected_min_raw)
                selected_max = parse_decimal(selected_max_raw)
                if selected_min is not None:
                    matching_ids = ProductAttributeValue.objects.filter(
                        id_product_id__in=products_qs.values('id_product'),
                        id_attribute=attribute,
                        value_number__gte=selected_min,
                    ).values('id_product')
                    products_qs = products_qs.filter(id_product__in=matching_ids)
                if selected_max is not None:
                    matching_ids = ProductAttributeValue.objects.filter(
                        id_product_id__in=products_qs.values('id_product'),
                        id_attribute=attribute,
                        value_number__lte=selected_max,
                    ).values('id_product')
                    products_qs = products_qs.filter(id_product__in=matching_ids)
                filter_item['min_key'] = min_key
                filter_item['max_key'] = max_key
                filter_item['min_value'] = number_bounds['min_value']
                filter_item['max_value'] = number_bounds['max_value']
                filter_item['selected_min'] = selected_min_raw
                filter_item['selected_max'] = selected_max_raw
                selected_attribute_filters[min_key] = selected_min_raw
                selected_attribute_filters[max_key] = selected_max_raw

            elif attribute.value_type == AttributeDefinition.VALUE_TYPE_BOOLEAN:
                bool_values = list(
                    ProductAttributeValue.objects.filter(
                        id_product_id__in=base_product_ids,
                        id_attribute=attribute,
                        value_boolean__isnull=False,
                    ).values_list('value_boolean', flat=True).distinct()
                )
                if not bool_values:
                    continue
                bool_key = f'attr_{attribute.id_attribute}'
                selected_bool_raw = request.GET.get(bool_key, '').strip()
                if selected_bool_raw in ('0', '1'):
                    selected_bool = selected_bool_raw == '1'
                    matching_ids = ProductAttributeValue.objects.filter(
                        id_product_id__in=products_qs.values('id_product'),
                        id_attribute=attribute,
                        value_boolean=selected_bool,
                    ).values('id_product')
                    products_qs = products_qs.filter(id_product__in=matching_ids)
                filter_item['param_key'] = bool_key
                filter_item['selected'] = selected_bool_raw
                filter_item['bool_values'] = sorted(set(int(bool(value)) for value in bool_values))
                selected_attribute_filters[bool_key] = selected_bool_raw
            else:
                continue

            category_filters.append(filter_item)

        products = products_qs.distinct()

    category_chain = CategoryService.build_category_chain(category)
    
    user = UserService.get_user_from_session(request)
    cart_item_count = CartService.get_cart_item_count(user)
    
    return render(request, 'shop/category_detail.html', {
        'category': category,
        'children': children,
        'products': products,
        'category_filters': category_filters,
        'price_min': min_price_raw or (str(price_bounds_min) if price_bounds_max else ''),
        'price_max': max_price_raw or (str(price_bounds_max) if price_bounds_max else ''),
        'price_bounds_min': price_bounds_min,
        'price_bounds_max': price_bounds_max,
        'selected_attribute_filters': selected_attribute_filters,
        'category_chain': category_chain,
        'cart_item_count': cart_item_count,
        'current_user': user,
    })

def product_detail(request, product_id):
    product = ProductService.get_product_by_id(product_id)
    category_chain = CategoryService.build_category_chain(product.id_category)
    reviews = ProductService.get_product_reviews(product)
    
    user = UserService.get_user_from_session(request)
    cart_item_count = CartService.get_cart_item_count(user)
    cart_item = CartService.get_cart_item_for_product(user, product)
    product_cart_quantity = cart_item.quantity if cart_item else 0
    can_leave_review, has_delivered_purchase, already_reviewed = get_review_access(user, product)
    reviews_count = len(reviews)

    avg_rating = reviews.aggregate(avg=Avg('rating')).get('avg')
    if avg_rating is None:
        avg_rating_display = None
        rounded_rating = 0
    else:
        avg_rating_display = f'{avg_rating:.1f}'.replace('.', ',')
        rounded_rating = max(1, min(5, int(avg_rating + 0.5)))
    
    parent_category_ids = set(
        Category.objects.filter(id_parent_category__isnull=False)
        .values_list('id_parent_category_id', flat=True)
    )
    leaf_categories = Category.objects.exclude(id_category__in=parent_category_ids).order_by('category_name')
    attribute_options_map = defaultdict(list)
    for option in AttributeOption.objects.filter(is_active=True).order_by('sort_order', 'option_value'):
        attribute_options_map[option.id_attribute_id].append({
            'id': option.id_option,
            'value': option.option_value,
        })
    category_attribute_config = defaultdict(list)
    mappings = (
        CategoryAttributeMap.objects.select_related('id_category', 'id_attribute')
        .filter(id_category_id__in=leaf_categories.values('id_category'))
        .order_by('id_category_id', 'step_order', 'id_category_attribute_map')
    )
    for mapping in mappings:
        attribute = mapping.id_attribute
        if not attribute or not attribute.is_active:
            continue
        category_attribute_config[mapping.id_category_id].append({
            'id': attribute.id_attribute,
            'name': attribute.attribute_name,
            'code': attribute.attribute_code,
            'value_type': attribute.value_type,
            'unit': attribute.unit or '',
            'options': attribute_options_map.get(attribute.id_attribute, []),
        })
    product_attribute_values = {
        pav.id_attribute_id: {
            'option_id': pav.id_option_id or '',
            'value_number': pav.value_number,
            'value_string': pav.value_string or '',
            'value_boolean': '' if pav.value_boolean is None else ('1' if pav.value_boolean else '0'),
        }
        for pav in ProductAttributeValue.objects.filter(id_product=product)
    }

    return render(request, 'shop/product_detail.html', {
        'product': product,
        'category_chain': category_chain,
        'reviews': reviews,
        'reviews_label': get_reviews_label(reviews_count),
        'avg_rating_display': avg_rating_display,
        'rounded_rating': rounded_rating,
        'cart_item_count': cart_item_count,
        'current_user': user,
        'product_cart_quantity': product_cart_quantity,
        'can_leave_review': can_leave_review,
        'has_delivered_purchase': has_delivered_purchase,
        'already_reviewed': already_reviewed,
        'can_manage_product': can_manage_store(user),
        'all_categories': leaf_categories if can_manage_store(user) else [],
        'category_attribute_config': dict(category_attribute_config),
        'product_attribute_values': product_attribute_values,
    })
